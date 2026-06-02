"""
convertidor2.py – Extractor Planilla Resumen (Aportes en Línea)
Funciona con cualquier PDF (Diciembre, Octubre, etc.)
- Extrae cada página con su tabla nativa
- Comprime cada fila quitando celdas vacías → estructura consistente
- Filtra solo filas que empiecen con número consecutivo
"""

import os
import pdfplumber
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog, messagebox


def extraer(ruta):
    todas = []

    with pdfplumber.open(ruta) as pdf:
        for pag in pdf.pages:
            tablas = pag.find_tables({
                "vertical_strategy":   "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 4,
            })
            if not tablas:
                continue

            # Tomar la tabla más grande de la página
            tabla = max(tablas, key=lambda t: len(t.extract())).extract()

            for fila in tabla:
                # Limpiar y comprimir (quitar celdas vacías)
                celdas = [str(c or "").replace("\n", " ").strip() for c in fila]
                compact = [v for v in celdas if v]

                if not compact:
                    continue

                # Solo filas que empiecen con número consecutivo
                primer = compact[0]
                if primer.isdigit() and 1 <= int(primer) <= 9999:
                    todas.append(compact)

    if not todas:
        return pd.DataFrame()

    # Normalizar ancho (todas las filas al mismo largo)
    ancho = max(len(f) for f in todas)
    todas = [f + [""] * (ancho - len(f)) for f in todas]

    df = pd.DataFrame(todas)
    df = df.sort_values(0, key=lambda s: pd.to_numeric(s, errors="coerce").fillna(9999))
    df = df.reset_index(drop=True)
    return df


def guardar(df, ruta_pdf):
    carpeta_datos = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Datos")
    os.makedirs(carpeta_datos, exist_ok=True)
    nombre_base = os.path.splitext(os.path.basename(ruta_pdf))[0]
    salida = os.path.join(carpeta_datos, f"{nombre_base}_PLANILLA.xlsx")

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Planilla")
        ws = writer.book["Planilla"]

        for ci in range(1, ws.max_column + 1):
            c = ws.cell(1, ci)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(horizontal="center")

        for ci in range(1, ws.max_column + 1):
            letra = get_column_letter(ci)
            ancho = max(len(str(ws.cell(r, ci).value or ""))
                        for r in range(1, ws.max_row + 1))
            ws.column_dimensions[letra].width = min(max(ancho + 2, 5), 28)

        ws.freeze_panes = "A2"

    return salida


if __name__ == "__main__":
    root = Tk()
    root.withdraw()
    archivo = filedialog.askopenfilename(
        title="Seleccionar PDF",
        filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")],
    )
    if not archivo:
        print("Cancelado.")
    else:
        try:
            df = extraer(archivo)
            if df.empty:
                messagebox.showerror("Error", "No se encontraron filas.")
            else:
                ruta = guardar(df, archivo)
                messagebox.showinfo("Listo", f"{len(df)} filas exportadas.\n{ruta}")
        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror("Error", str(e))
