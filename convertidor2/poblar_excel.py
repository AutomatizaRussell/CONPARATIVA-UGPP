"""
Toma los JSON de la carpeta Datos (estructurado + facturas del mismo mes)
y genera un Excel consolidado por persona.

  PAGO TRABAJADOR Y SS (facturas JSON):
  A - TIPO             <- estructurado  "IDENTIFICA CIÓN - Tipo"
  B - NÚMERO           <- estructurado  "documento"
  C - NOMBRE           <- estructurado  "NOMBRES"
  D - SALARIO MES      <- facturas  códigos 001+051+052  (Valor Unitario)
  E - BONO NO SALARIO  <- facturas  códigos 022+066      (Valor Unitario)
  F - SS BONO...       <- (pendiente)
  G - AUX. ROD.        <- facturas  código  026          (Valor Unitario)
  H - PENSIÓN trab.    <- facturas  códigos 700-798      (Valor Unitario)
  I - SALUD trab.      <- facturas  códigos 600-699      (Valor Unitario)
  J - FSP trab.        <- facturas  código  799          (Valor Unitario)

  PAGO TOTAL SS EMPRESA (PLANILLA JSON - claves numéricas):
  K - PENSIÓN emp.     <- clave "7" - (clave "6" * 1%)   (AFP Aporte menos FSP)
  L - FSP emp.         <- clave "6" * 1%                  (1% del IBC AFP)
  M - SALUD emp.       <- clave "11"                      (EPS Aporte)
  N - ARL              <- clave "19"                      (ARL Aporte)
  O - CCF              <- clave "15"                      (CCF Aporte)
  P - SENA             <- clave "22" * 40%                (40% Parafiscales)
  Q - ICBF             <- clave "22" * 60%                (60% Parafiscales)
"""

import os
import re
import json
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from excel import crear_excel

CARPETA_DATOS = Path(__file__).parent / "Datos"

CODIGOS_SALARIO  = {"001", "051", "052"}
CODIGO_BONO      = {"022", "066"}
CODIGO_AUX_ROD   = "026"
RANGO_PENSION    = (700, 798)
RANGO_SALUD      = (600, 699)
CODIGO_FSP       = "799"


# ── helpers ──────────────────────────────────────────────────────────────────

def parse_valor(texto: str) -> float:
    """Facturas: '3,733,333.00' → 3733333.0  (coma=miles, punto=decimal)"""
    limpio = re.sub(r"[^\d.,]", "", str(texto))
    if "," in limpio and "." in limpio:
        limpio = limpio.replace(",", "")
    elif "," in limpio:
        limpio = limpio.replace(",", ".")
    try:
        return float(limpio)
    except ValueError:
        return 0.0


def parse_valor_col(texto: str) -> float:
    """Estructurado: '$ 45.400' → 45400.0  (punto=miles, coma=decimal, formato colombiano)"""
    limpio = re.sub(r"[^\d.,]", "", str(texto))
    if not limpio:
        return 0.0
    if "," in limpio:
        # coma como decimal: quitar puntos de miles, convertir coma
        limpio = limpio.replace(".", "").replace(",", ".")
    else:
        # solo puntos → separador de miles, no hay decimales
        limpio = limpio.replace(".", "")
    try:
        return float(limpio)
    except ValueError:
        return 0.0


def parse_planilla(texto: str) -> float:
    """Planilla: '$8,000,000' → 8000000.0  ($ + comas como miles)"""
    limpio = re.sub(r"[^\d]", "", str(texto))
    try:
        return float(limpio) if limpio else 0.0
    except ValueError:
        return 0.0


def extraer_est(registros: list) -> dict:
    """
    Suma los valores SS empresa desde los registros del PLANILLA JSON.
    Claves numéricas: "6"=AFP IBC, "7"=AFP Aporte, "11"=EPS Aporte,
                      "15"=CCF Aporte, "19"=ARL Aporte, "22"=Para Aporte
    """
    totales = {"pension_emp": 0.0, "fsp_emp": 0.0, "salud_emp": 0.0,
               "arl": 0.0, "ccf": 0.0, "sena": 0.0, "icbf": 0.0}
    for r in registros:
        afp_ibc    = parse_planilla(r.get("6", "0"))
        afp_aporte = parse_planilla(r.get("7", "0"))
        totales["pension_emp"] += afp_aporte - (afp_ibc * 0.01)
        totales["fsp_emp"]     += afp_ibc * 0.01
        totales["salud_emp"]   += parse_planilla(r.get("11", "0"))
        totales["arl"]         += parse_planilla(r.get("19", "0"))
        totales["ccf"]         += parse_planilla(r.get("15", "0"))
        para = parse_planilla(r.get("22", "0"))
        totales["sena"]        += para * 0.40
        totales["icbf"]        += para * 0.60
    return totales


def aplanar_estructurado(data: dict) -> dict:
    """
    Devuelve { documento: [lista_de_todos_los_registros] }.
    Acumula registros de todas las Tabla_N que compartan el mismo documento.
    """
    plano: dict[str, list] = {}
    primera_clave = next(iter(data), "")
    iterables = data.values() if primera_clave.startswith("Tabla_") else [data]
    for bloque in iterables:
        for doc, info in bloque.items():
            if doc not in plano:
                plano[doc] = []
            plano[doc].extend(info.get("registros", []))
    return plano


def _sum_codigos(registros: list, codigos: set) -> float:
    return sum(
        parse_valor(r.get("Valor Unitario", "0"))
        for r in registros
        if str(r.get("Código", "")).strip() in codigos
    )


def _sum_rango(registros: list, rango: tuple) -> float:
    lo, hi = rango
    total = 0.0
    for r in registros:
        try:
            cod = int(str(r.get("Código", "")).strip())
        except ValueError:
            continue
        if lo <= cod <= hi:
            total += parse_valor(r.get("Valor Unitario", "0"))
    return total


def extraer_columnas(registros: list) -> dict:
    """Calcula todos los valores monetarios de una persona desde sus registros de facturas."""
    return {
        "salario":  _sum_codigos(registros, CODIGOS_SALARIO),
        "bono":     _sum_codigos(registros, CODIGO_BONO),
        "aux_rod":  _sum_codigos(registros, {CODIGO_AUX_ROD}),
        "pension":  _sum_rango(registros, RANGO_PENSION),
        "salud":    _sum_rango(registros, RANGO_SALUD),
        "fsp":      _sum_codigos(registros, {CODIGO_FSP}),
    }


def detectar_pares() -> list[tuple[str, Path, Path]]:
    """
    Busca pares  *_PLANILLA.json + *_facturas.json con el mismo prefijo.
    Devuelve lista de (prefijo, ruta_planilla, ruta_facturas).
    """
    planillas = {f.stem.replace("_PLANILLA", ""): f
                 for f in CARPETA_DATOS.glob("*_PLANILLA.json")}
    facturas  = {f.stem.replace("_facturas", ""): f
                 for f in CARPETA_DATOS.glob("*_facturas.json")}

    pares = []
    for prefijo in sorted(planillas):
        if prefijo in facturas:
            pares.append((prefijo, planillas[prefijo], facturas[prefijo]))
        else:
            print(f"  Aviso: '{prefijo}_PLANILLA.json' sin facturas — se omite.")
    return pares


# ── formato ──────────────────────────────────────────────────────────────────

def aplicar_formato_datos(ws, fila_inicio: int, fila_fin: int):
    thin = Side(border_style="thin", color="D3D3D3")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_par = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    font = Font(name="Arial", size=10)
    centro = Alignment(horizontal="center", vertical="center")
    derecha = Alignment(horizontal="right", vertical="center")
    izq = Alignment(horizontal="left", vertical="center")

    for row_idx in range(fila_inicio, fila_fin + 1):
        ws.row_dimensions[row_idx].height = 18
        es_par = (row_idx % 2 == 0)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font
            cell.border = borde
            if es_par:
                cell.fill = fill_par
            # Alineación por columna
            if col_idx in (1, 2):                        # TIPO, NÚMERO → centrado texto
                cell.alignment = centro
                cell.number_format = "@"
            elif col_idx in (4, 5, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17):  # monetarias
                cell.alignment = derecha
                cell.number_format = '#,##0'
            else:
                cell.alignment = izq


# ── lógica principal ──────────────────────────────────────────────────────────

def procesar_par(prefijo: str, ruta_est: Path, ruta_fac: Path):
    print(f"\nProcesando: {prefijo}")

    with open(ruta_est, encoding="utf-8") as f:
        data_est = json.load(f)
    with open(ruta_fac, encoding="utf-8") as f:
        data_fac = json.load(f)

    personas_est = aplanar_estructurado(data_est)   # { doc: primer_registro }
    personas_fac = data_fac                          # { doc: { registros: [...] } }

    # Todos los documentos de facturas (base principal)
    # Los que también estén en estructurado tendrán TIPO y NOMBRES de allí
    todos_docs = sorted(personas_fac)
    if not todos_docs:
        print(f"  Sin documentos en facturas — se omite.")
        return

    # Crear Excel base
    ruta_excel = CARPETA_DATOS / f"{prefijo}_CONSOLIDADO.xlsx"
    crear_excel(destino=str(ruta_excel))

    wb = openpyxl.load_workbook(ruta_excel)
    ws = wb.active
    fila = 3  # Los datos empiezan en fila 3 (1=grupos, 2=encabezados)

    for doc in todos_docs:
        reg_fac  = personas_fac[doc]["registros"]
        regs_est = personas_est.get(doc, [])       # lista; puede estar vacía

        # A/B/C: del primer registro del planilla (claves numéricas "1","2","3")
        primer_est = regs_est[0] if regs_est else {}
        primer_fac = reg_fac[0]  if reg_fac  else {}
        tipo   = str(primer_est.get("1", "")).strip() or "CC"
        numero = str(primer_est.get("2", doc)).strip()
        nombre = (str(primer_est.get("3", "")).strip()
                  or str(primer_fac.get("Nombre Empleado", "")).strip())
        vals = extraer_columnas(reg_fac)
        emp  = extraer_est(regs_est)               # suma todos los registros

        ws.cell(row=fila, column=1).value  = tipo
        ws.cell(row=fila, column=2).value  = numero
        ws.cell(row=fila, column=3).value  = nombre
        ws.cell(row=fila, column=4).value  = vals["salario"]  or None
        ws.cell(row=fila, column=5).value  = vals["bono"]     or None
        # columna 6 (F - SS BONO NO SALARIAL) pendiente
        ws.cell(row=fila, column=7).value  = vals["aux_rod"]  or None
        ws.cell(row=fila, column=8).value  = vals["pension"]  or None
        ws.cell(row=fila, column=9).value  = vals["salud"]    or None
        ws.cell(row=fila, column=10).value = vals["fsp"]      or None
        ws.cell(row=fila, column=11).value = emp["pension_emp"] or None
        ws.cell(row=fila, column=12).value = emp["fsp_emp"]   or None
        ws.cell(row=fila, column=13).value = emp["salud_emp"] or None
        ws.cell(row=fila, column=14).value = emp["arl"]       or None
        ws.cell(row=fila, column=15).value = emp["ccf"]       or None
        ws.cell(row=fila, column=16).value = emp["sena"]      or None
        ws.cell(row=fila, column=17).value = emp["icbf"]      or None

        fila += 1

    aplicar_formato_datos(ws, fila_inicio=3, fila_fin=fila - 1)

    wb.save(ruta_excel)
    print(f"  -> {ruta_excel.name}  ({fila - 3} personas)")


def main():
    if not CARPETA_DATOS.exists():
        print(f"La carpeta Datos no existe: {CARPETA_DATOS}")
        return

    pares = detectar_pares()
    if not pares:
        print("No se encontraron pares estructurado+facturas en la carpeta Datos.")
        return

    for prefijo, ruta_est, ruta_fac in pares:
        procesar_par(prefijo, ruta_est, ruta_fac)

    print("\nListo.")


if __name__ == "__main__":
    main()
