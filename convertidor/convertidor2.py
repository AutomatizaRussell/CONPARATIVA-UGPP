"""
convertidor2.py – Extractor Planilla Resumen (Aportes en Línea)
- Detecta las columnas desde la tabla real del PDF
- Fuerza esas mismas columnas en TODAS las páginas
- Filtra solo filas que empiecen con número consecutivo
"""

import os
import pdfplumber
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog, messagebox


def detectar_col_xs(pdf):
    """
    Extrae las posiciones X de columnas desde la tabla con MENOS columnas
    que tenga filas numeradas (evita tomar página 1 que incluye encabezado).
    """
    candidatos = []
    for pag in pdf.pages:
        tablas = pag.find_tables({
            "vertical_strategy":   "lines",
            "horizontal_strategy": "lines",
            "snap_tolerance": 4,
        })
        for t in tablas:
            data = t.extract()
            if any(str(f[0] or "").strip().isdigit() for f in data if f):
                if t.cells:
                    xs = sorted(set(round(c[0], 1) for c in t.cells))
                    if len(xs) >= 10:
                        candidatos.append(xs)
    if not candidatos:
        return None
    # Usar la que tenga MENOS posiciones X (tabla más limpia, sin encabezado)
    xs = min(candidatos, key=len)
    # Añadir borde derecho de la tabla para capturar la última columna (Total)
    for pag in pdf.pages:
        tablas = pag.find_tables({
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "snap_tolerance": 4,
        })
        for t in tablas:
            data = t.extract()
            if len(data[0]) == len(xs) if data else False:
                borde_der = round(t.bbox[2], 1)
                if borde_der not in xs:
                    xs = xs + [borde_der]
                return xs
    return xs


def extraer(ruta):
    todas = []

    with pdfplumber.open(ruta) as pdf:

        # Obtener posiciones X de columnas (referencia desde cualquier página)
        col_xs = detectar_col_xs(pdf)
        if not col_xs:
            raise RuntimeError("No se encontró la tabla en el PDF.")

        # Configuración que FUERZA las mismas columnas en todas las páginas
        settings = {
            "vertical_strategy":      "explicit",
            "explicit_vertical_lines": col_xs,
            "horizontal_strategy":    "lines",
            "snap_tolerance":          4,
            "join_tolerance":          4,
        }

        for pag in pdf.pages:
            tabla = pag.extract_table(settings)
            if not tabla:
                continue

            for fila in tabla:
                celdas = [str(c or "").replace("\n", " ").strip() for c in fila]
                primer = celdas[0] if celdas else ""
                if primer.isdigit() and 1 <= int(primer) <= 9999:
                    todas.append(celdas)

    if not todas:
        return pd.DataFrame()

    ancho = max(len(f) for f in todas)
    todas = [f + [""] * (ancho - len(f)) for f in todas]

    df = pd.DataFrame(todas)
    df = df.loc[:, (df != "").any(axis=0)]
    df = df.sort_values(0, key=lambda s: pd.to_numeric(s, errors="coerce").fillna(9999))
    df = df.reset_index(drop=True)
    return df


def guardar(df, ruta_pdf):
    carpeta_datos = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Datos")
    os.makedirs(carpeta_datos, exist_ok=True)
    nombre_base = os.path.splitext(os.path.basename(ruta_pdf))[0]
    salida = os.path.join(carpeta_datos, f"{nombre_base}_PLANILLA.xlsx")

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Planilla")
        ws = writer.book["Planilla"]

        for ci in range(1, ws.max_column + 1):
            c = ws.cell(1, ci)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(horizontal="center")

        for ci in range(1, ws.max_column + 1):
            letra = get_column_letter(ci)
            ancho = max(len(str(ws.cell(r, ci).value or ""))
                        for r in range(1, ws.max_row + 1))
            ws.column_dimensions[letra].width = min(max(ancho + 2, 5), 28)

        ws.freeze_panes = "A2"

    return salida


if __name__ == "__main__":
    root = Tk()
    root.withdraw()
    archivo = filedialog.askopenfilename(
        title="Seleccionar PDF",
        filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")],
    )
    if not archivo:
        print("Cancelado.")
    else:
        try:
            df = extraer(archivo)
            if df.empty:
                messagebox.showerror("Error", "No se encontraron filas.")
            else:
                ruta = guardar(df, archivo)
                messagebox.showinfo("Listo", f"{len(df)} filas exportadas.\n{ruta}")
        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror("Error", str(e))
