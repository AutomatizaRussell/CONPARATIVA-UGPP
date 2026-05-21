import pdfplumber
import pandas as pd
import re
from tkinter import Tk, filedialog, messagebox
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def separar_documento(texto):
    """Separa tipos de documento (CC, CE, etc.) de los números."""
    if not texto:
        return "", ""
    texto = str(texto).strip()
    patron = r'^([cC]\.?[cC]\.?|[cC]\.?[eE]\.?|[nN]\.?[iI]\.?[tT]\.?|[tT]\.?[iI]\.?|[pP][aA][sS][aA][pP][oO][rR][tT][eE]?|[rR][cC])\s*[-_]*\s*(.*)$'
    match = re.match(patron, texto)
    if match:
        tipo = match.group(1).upper().replace(".", "").strip()
        numero = match.group(2).strip()
        return tipo, numero
    return "", texto

def aplicar_formato_excel(ruta_excel):
    """Abre el Excel generado y le aplica un formato estético e inteligente."""
    wb = openpyxl.load_workbook(ruta_excel)
    
    # Paleta de colores profesionales (Azul Ejecutivo Desaturado)
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10, bold=False, color="000000")
    
    fill_header = PatternFill(start_color="2A4D69", end_color="2A4D69", fill_type="solid")
    fill_zebra = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    
    # Bordes finos gris claro para simular cuadrícula de cuaderno estructurado
    thin_side = Side(border_style="thin", color="D3D3D3")
    border_celda = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True # Asegura que las líneas guía se vean
        ws.row_dimensions[1].height = 26 # Encabezado con más aire
        
        # 1. Dar estilo al encabezado (Fila 1)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_celda
            
        # 2. Dar estilo y formato a los datos
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20 # Filas de datos estilizadas
            es_par = (row_idx % 2 == 0)
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_body
                cell.border = border_celda
                
                # Efecto Cebra intercalado
                if es_par:
                    cell.fill = fill_zebra
                
                # Alineación inteligente según el tipo de dato
                valor_str = str(cell.value).strip() if cell.value is not None else ""
                
                # Si es la columna de número de documento, forzar formato texto para no perder ceros a la izquierda
                if "NÚMERO" in str(ws.cell(row=1, column=col_idx).value).upper():
                    cell.alignment = align_left
                    cell.number_format = '@'
                # Si es un número consecutivo (Nº) o códigos cortos, centrar
                elif col_idx == 1 or len(valor_str) <= 3 or valor_str.upper() in ["CC", "CE", "NIT", "TI", "RC"]:
                    cell.alignment = align_center
                # Si parece un valor monetario o número grande, alinear a la derecha
                elif valor_str.replace('.', '', 1).isdigit() and len(valor_str) > 5 and not "-" in valor_str:
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
        
        # 3. Auto-ajustar el ancho de las columnas para que nada se corte (###)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Margen de seguridad para que no se pegue al borde
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(ruta_excel)

def seleccionar_y_convertir():
    root = Tk()
    root.withdraw()
    
    print("Esperando selección de archivo...")
    archivo_pdf = filedialog.askopenfilename(
        title="Selecciona el archivo PDF para segmentar",
        filetypes=[("Archivos PDF", "*.pdf")]
    )
    
    if not archivo_pdf:
        print("Conversión cancelada.")
        return

    carpeta_datos = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Datos")
    os.makedirs(carpeta_datos, exist_ok=True)
    nombre_base = os.path.splitext(os.path.basename(archivo_pdf))[0]
    archivo_excel = os.path.join(carpeta_datos, f"{nombre_base}_estructurado.xlsx")

    todas_las_tablas = []
    tabla_actual = []
    encabezados_actuales = None
    buscando_tabla = False

    print(f"Analizando estructura en: {archivo_pdf}")

    try:
        with pdfplumber.open(archivo_pdf) as pdf:
            for num_pagina, pagina in enumerate(pdf.pages):
                
                filas_pagina = pagina.extract_table(table_settings={
                    "vertical_strategy": "lines", 
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 3,
                    "join_tolerance": 3
                })
                
                if not filas_pagina:
                    continue
                
                for fila in filas_pagina:
                    fila_limpia = [str(celda).replace('\n', ' ').strip() if celda is not None else "" for celda in fila]
                    
                    # Detectar Inicio por "Nº"
                    if fila_limpia and ("Nº" in fila_limpia[0] or "N°" in fila_limpia[0] or fila_limpia[0] == "N"):
                        if buscando_tabla and tabla_actual:
                            df_temp = pd.DataFrame(tabla_actual, columns=encabezados_actuales)
                            todas_las_tablas.append(df_temp)
                            tabla_actual = []
                        
                        encabezados_actuales = fila_limpia
                        buscando_tabla = True
                        continue
                    
                    # Detectar Fin por "DATOS DEL COTIZANTE"
                    if buscando_tabla and any("DATOS DEL COTIZANTE" in str(celda).upper() for celda in fila_limpia):
                        if tabla_actual:
                            df_temp = pd.DataFrame(tabla_actual, columns=encabezados_actuales)
                            todas_las_tablas.append(df_temp)
                            tabla_actual = []
                        buscando_tabla = False
                        continue
                    
                    if buscando_tabla:
                        if len(fila_limpia) == len(encabezados_actuales):
                            tabla_actual.append(fila_limpia)
                        elif len(fila_limpia) < len(encabezados_actuales):
                            fila_limpia += [""] * (len(encabezados_actuales) - len(fila_limpia))
                            tabla_actual.append(fila_limpia[:len(encabezados_actuales)])

        if buscando_tabla and tabla_actual:
            df_temp = pd.DataFrame(tabla_actual, columns=encabezados_actuales)
            todas_las_tablas.append(df_temp)

        # Escribir y dar formato estético
        if todas_las_tablas:
            with pd.ExcelWriter(archivo_excel, engine='openpyxl') as writer:
                for i, df in enumerate(todas_las_tablas):
                    if df.empty:
                        continue
                        
                    # Aplicar la separación de documentos CC / CE
                    for col in df.columns:
                        tiene_documentos = df[col].astype(str).str.contains(r'CC|CE|NIT|TI|cc|ce|nit|ti', regex=True).any()
                        if tiene_documentos:
                            tipos, numeros = zip(*df[col].apply(separar_documento))
                            idx = df.columns.get_loc(col)
                            df.insert(idx, f"{col} - Tipo", tipos)
                            df.insert(idx + 1, f"{col} - Número", numeros)
                            df.drop(columns=[col], inplace=True)
                            break 
                    
                    df.to_excel(writer, sheet_name=f"Tabla_{i+1}", index=False)
            
            # --- EMBELLECER EL EXCEL ---
            aplicar_formato_excel(archivo_excel)
            
            print(f"¡Éxito total! Archivo estilizado en: {archivo_excel}")
            messagebox.showinfo("¡Éxito!", f"Se procesaron {len(todas_las_tablas)} tablas.\n\nEl archivo se guardó con diseño de tabla profesional, fuentes estandarizadas, efecto cebra y autoajuste de celdas.")
        else:
            messagebox.showwarning("Aviso", "No se encontraron los patrones de inicio/fin en las tablas.")

    except Exception as e:
        print(f"Error: {e}")
        messagebox.showerror("Error", f"Ocurrió un error en el motor de diseño:\n{e}")

if __name__ == "__main__":
    seleccionar_y_convertir()