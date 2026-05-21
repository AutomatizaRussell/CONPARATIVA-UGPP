import pdfplumber
import pandas as pd
import re
from tkinter import Tk, filedialog, messagebox
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def aplicar_formato_facturas(ruta_excel):
    """Aplica diseño corporativo al Excel de facturas."""
    wb = openpyxl.load_workbook(ruta_excel)
    
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10, bold=False, color="000000")
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    border_celda = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        ws.row_dimensions[1].height = 26
        
        # Estilo Encabezados
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_celda
            
        # Estilo Datos
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            es_par = (row_idx % 2 == 0)
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_body
                cell.border = border_celda
                if es_par:
                    cell.fill = fill_zebra
                
                col_name = str(ws.cell(row=1, column=col_idx).value).upper()
                if any(x in col_name for x in ["CÉDULA", "TURNO", "CÓDIGO"]):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = '@'
                elif any(x in col_name for x in ["VALOR", "CANTIDAD", "TOTAL"]):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(ruta_excel)

def seleccionar_y_convertir_facturas():
    root = Tk()
    root.withdraw()
    
    print("Esperando selección de PDF de facturas...")
    archivo_pdf = filedialog.askopenfilename(
        title="Selecciona el PDF de la Factura",
        filetypes=[("Archivos PDF", "*.pdf")]
    )
    
    if not archivo_pdf:
        print("Conversión cancelada.")
        return

    # --- NORMALIZACIÓN ABSOLUTA DE RUTAS PARA WINDOWS Y ONEDRIVE ---
    archivo_pdf = os.path.abspath(os.path.normpath(archivo_pdf))
    carpeta_datos = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Datos")
    os.makedirs(carpeta_datos, exist_ok=True)
    nombre_base = os.path.splitext(os.path.basename(archivo_pdf))[0]

    archivo_excel = os.path.abspath(os.path.join(carpeta_datos, f"{nombre_base}_facturas.xlsx"))
    archivo_debug = os.path.abspath(os.path.join(carpeta_datos, f"{nombre_base}_TEXTO_CRUDO.txt"))

    datos_factura_total = []
    
    cedula_actual = "No Detectado"
    nombre_actual = "No Detectado"
    turno_actual = "No Detectado"
    texto_crudo_completo = ""

    print(f"Procesando: {archivo_pdf}")

    try:
        with pdfplumber.open(archivo_pdf) as pdf:
            for pagina in pdf.pages:
                texto_pagina = pagina.extract_text(layout=True)
                if not texto_pagina:
                    texto_pagina = pagina.extract_text()
                if not texto_pagina:
                    continue
                
                texto_crudo_completo += texto_pagina + "\n"
                lineas = texto_pagina.split("\n")
                
                for linea in lineas:
                    linea_str = linea.strip()
                    if not linea_str:
                        continue
                    
                    if "TURNO" in linea_str.upper():
                        match_cedula = re.search(r'^(\d{5,})', linea_str)
                        match_turno = re.search(r'Turno\s*[:\-]*\s*(\w+)', linea_str, re.IGNORECASE)
                        
                        if match_cedula:
                            cedula_actual = match_cedula.group(1)
                            nombre_bloque = linea_str[len(cedula_actual):]
                            idx_turno = nombre_bloque.upper().find("TURNO")
                            if idx_turno != -1:
                                nombre_actual = nombre_bloque[:idx_turno].strip()
                            else:
                                nombre_actual = nombre_bloque.strip()
                        
                        if match_turno:
                            turno_actual = match_turno.group(1)
                        continue
                    
                    match_concepto_inicio = re.match(r'^(\d{3})\s+(.*)$', linea_str)
                    if match_concepto_inicio:
                        cod = match_concepto_inicio.group(1)
                        resto_linea = match_concepto_inicio.group(2).strip()
                        
                        partes_valores = [p.strip() for p in re.split(r'\s{2,}', resto_linea) if p.strip()]
                        
                        concepto = partes_valores[0] if len(partes_valores) > 0 else ""
                        col_cantidad = partes_valores[1] if len(partes_valores) > 1 else ""
                        col_valor = partes_valores[2] if len(partes_valores) > 2 else ""
                        col_total = partes_valores[3] if len(partes_valores) > 3 else ""
                        
                        datos_factura_total.append({
                            "Cédula": cedula_actual,
                            "Nombre Empleado": nombre_actual,
                            "Turno": turno_actual,
                            "Código": cod,
                            "Concepto": concepto,
                            "Cantidad / Horas": col_cantidad,
                            "Valor Unitario": col_valor,
                            "Total": col_total
                        })

        # Guardar log de texto
        with open(archivo_debug, "w", encoding="utf-8") as f:
            f.write(texto_crudo_completo)

        # Generar Excel
        if datos_factura_total:
            df_final = pd.DataFrame(datos_factura_total)
            
            # Guardado explícito y cierre de flujo de datos
            writer = pd.ExcelWriter(archivo_excel, engine='openpyxl')
            df_final.to_excel(writer, index=False)
            writer.close()
            
            # Dar formato estético
            aplicar_formato_facturas(archivo_excel)
            
            # --- SOLUCIÓN DE VISIBILIDAD: FORZAR APERTURA ---
            print("Forzando apertura de archivos...")
            
            # Mensaje de éxito
            messagebox.showinfo("¡Éxito!", f"Excel generado.\n\nA continuación se abrirá el archivo de forma automática.")
            
            # Abrir el Excel directamente
            try:
                os.startfile(archivo_excel)
            except Exception as e_open:
                print(f"No se pudo abrir el archivo directamente, abriendo carpeta: {e_open}")
                os.startfile(carpeta_datos) # Si falla abrir el archivo, abre la carpeta
                
        else:
            messagebox.showwarning("Sin Registros", "No se detectaron los formatos de conceptos.")

    except Exception as e:
        messagebox.showerror("Error Crítico", f"Ocurrió un error inesperado:\n{e}")

if __name__ == "__main__":
    seleccionar_y_convertir_facturas()