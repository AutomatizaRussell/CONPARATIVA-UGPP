from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from pathlib import Path

DESTINO = Path(__file__).parent / "RESULTADO.xlsx"


def crear_excel(destino=None):
    ruta = Path(destino) if destino else DESTINO

    wb = Workbook()
    ws = wb.active
    ws.title = "ENERO"

    anchos = {
        "A": 4.54,  "B": 12.45, "C": 30.91, "D": 15.18,
        "E": 16.82, "F": 21.63, "G": 12.54, "H": 12.63,
        "I": 8.0,   "J": 13.82, "K": 14.09, "L": 12.63,
        "M": 14.09, "N": 11.54, "O": 12.54, "P": 12.54,
        "Q": 8.0,   "R": 37.82,
    }
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    bold = Font(bold=True, size=11)
    centro = Alignment(horizontal="center")
    borde_thin = Border(
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
    )

    ws.merge_cells("A1:C1")
    ws["A1"].alignment = centro

    ws.merge_cells("D1:J1")
    ws["D1"].value = "PAGO TRABAJADOR Y SS"
    ws["D1"].font = bold
    ws["D1"].alignment = centro

    ws.merge_cells("K1:O1")
    ws["K1"].value = "PAGO TOTAL SS EMPRESA"
    ws["K1"].font = bold
    ws["K1"].alignment = centro

    encabezados = [
        ("A2", "TIPO",                None),
        ("B2", "NÚMERO",              centro),
        ("C2", "NOMBRE",              centro),
        ("D2", "SALARIO MES",         centro),
        ("E2", "BONO NO SALARIO",     centro),
        ("F2", "SS BONO NO SALARIAL", centro),
        ("G2", "AUX. ROD.",           centro),
        ("H2", "PENSIÓN",             centro),
        ("I2", "SALUD",               centro),
        ("J2", "FSP",                 centro),
        ("K2", "PENSIÓN ",            centro),
        ("L2", "FSP",                 centro),
        ("M2", "SALUD",               centro),
        ("N2", "ARL",                 centro),
        ("O2", "CCF",                 centro),
        ("P2", "SENA",                centro),
        ("Q2", "ICBF",                centro),
        ("R2", "OBSERVACIONES",       centro),
    ]

    for coord, valor, alineacion in encabezados:
        celda = ws[coord]
        celda.value = valor
        celda.font = bold
        celda.border = borde_thin
        if alineacion:
            celda.alignment = alineacion

    wb.save(ruta)
    return ruta


if __name__ == "__main__":
    ruta = crear_excel()
    print(f"Archivo creado: {ruta}")
